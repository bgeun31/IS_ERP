import io
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models import Asset, DeviceSnapshot, LogFile, User

router = APIRouter(prefix="/api/assets", tags=["assets"])

# 자동 필드 키 → Asset 컬럼명 매핑
_AUTO_FIELD_MAP = {
    "manufacturer": "manufacturer_override",
    "model": "model_override",
    "serial_number": "serial_number_override",
    "hostname": "hostname_override",
    "os": "os_override",
    "ip": "ip_override",
}

_MANUAL_FIELDS = [
    "asset_number", "resource_status", "idc_name", "floor_name",
    "rack_row", "rack_name", "hole_number", "status_change_date",
    "device_category", "asset_inspection", "status_inspection",
    "config_inspection", "env_inspection", "telnet_accessible",
    "asset_sticker", "rfid_attached", "cmdb_match",
    "uplink_redundancy", "vim_module", "note_before_after", "note",
]

# 엑셀 헤더 → 필드 매핑
_EXCEL_HEADER_MAP = {
    "자산번호": "asset_number",
    "자원상태": "resource_status",
    "제조사": "manufacturer",
    "모델": "model",
    "일련번호": "serial_number",
    "IDC명": "idc_name",
    "상면명": "floor_name",
    "랙열명": "rack_row",
    "랙명": "rack_name",
    "홀번호": "hole_number",
    "상태변경일": "status_change_date",
    "호스트명": "hostname",
    "장비분류": "device_category",
    "OS": "os",
    "IP": "ip",
    "자산점검": "asset_inspection",
    "상태점검": "status_inspection",
    "설정점검": "config_inspection",
    "환경점검": "env_inspection",
    "텔넷접속가능여부": "telnet_accessible",
    "자산번호스티커": "asset_sticker",
    "RFID부착": "rfid_attached",
    "CMDB일치": "cmdb_match",
    "업링크이중화": "uplink_redundancy",
    "VIM모듈장착": "vim_module",
    "비고(수정전,후내용표기)": "note_before_after",
    "비고": "note",
}


class AssetUpdate(BaseModel):
    manufacturer: Optional[str] = None
    model: Optional[str] = None
    serial_number: Optional[str] = None
    hostname: Optional[str] = None
    os: Optional[str] = None
    ip: Optional[str] = None
    asset_number: Optional[str] = None
    resource_status: Optional[str] = None
    idc_name: Optional[str] = None
    floor_name: Optional[str] = None
    rack_row: Optional[str] = None
    rack_name: Optional[str] = None
    hole_number: Optional[str] = None
    status_change_date: Optional[str] = None
    device_category: Optional[str] = None
    asset_inspection: Optional[str] = None
    status_inspection: Optional[str] = None
    config_inspection: Optional[str] = None
    env_inspection: Optional[str] = None
    telnet_accessible: Optional[str] = None
    asset_sticker: Optional[str] = None
    rfid_attached: Optional[str] = None
    cmdb_match: Optional[str] = None
    uplink_redundancy: Optional[str] = None
    vim_module: Optional[str] = None
    note_before_after: Optional[str] = None
    note: Optional[str] = None


def _asset_row(asset: Asset) -> dict:
    """Asset 레코드를 응답 dict로 변환."""
    row: dict = {"device_name": asset.device_name}
    for key, col in _AUTO_FIELD_MAP.items():
        row[key] = getattr(asset, col, None)
    for field in _MANUAL_FIELDS:
        row[field] = getattr(asset, field, None)
    return row


def _snap_row(snap: DeviceSnapshot) -> dict:
    """DeviceSnapshot (Asset 미존재)을 응답 dict로 변환."""
    row: dict = {
        "device_name": snap.device_name,
        "manufacturer": snap.manufacturer,
        "model": snap.system_type,
        "serial_number": snap.serial_number,
        "hostname": snap.sysname or snap.device_name,
        "os": snap.primary_version,
        "ip": snap.management_ip,
    }
    for field in _MANUAL_FIELDS:
        row[field] = None
    return row


def _get_latest_snaps(db: Session) -> dict:
    """장비별 최신 스냅샷 dict 반환."""
    subq = (
        db.query(
            DeviceSnapshot.device_name,
            func.max(LogFile.log_year * 100 + LogFile.log_month).label("latest_ym"),
        )
        .join(LogFile, DeviceSnapshot.log_file_id == LogFile.id)
        .group_by(DeviceSnapshot.device_name)
        .subquery()
    )
    snaps = (
        db.query(DeviceSnapshot)
        .join(LogFile, DeviceSnapshot.log_file_id == LogFile.id)
        .join(
            subq,
            (DeviceSnapshot.device_name == subq.c.device_name)
            & ((LogFile.log_year * 100 + LogFile.log_month) == subq.c.latest_ym),
        )
        .all()
    )
    return {s.device_name: s for s in snaps}


@router.get("")
def list_assets(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """자산 목록 반환. Asset 레코드 우선, Asset 없는 로그 장비도 포함."""
    assets = db.query(Asset).all()
    deleted_names = {asset.device_name for asset in assets if asset.deleted}
    asset_map = {asset.device_name: asset for asset in assets if not asset.deleted}
    snap_map = {
        device_name: snap
        for device_name, snap in _get_latest_snaps(db).items()
        if device_name not in deleted_names
    }

    all_names = sorted(set(asset_map.keys()) | set(snap_map.keys()))
    result = []
    for name in all_names:
        asset = asset_map.get(name)
        if asset:
            result.append(_asset_row(asset))
        else:
            result.append(_snap_row(snap_map[name]))
    return result


# ── 정적 경로가 /{device_name} 보다 먼저 와야 함 ──


@router.post("/upload")
async def upload_asset_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """엑셀 파일에서 자산 데이터를 일괄 업로드합니다. 호스트명 열을 기준으로 매칭."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="파일이 없습니다")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="xlsx 또는 xls 파일만 지원합니다")

    data = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="시트를 찾을 수 없습니다")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일을 읽을 수 없습니다: {e}")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="데이터가 없습니다 (헤더 + 최소 1행 필요)")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    col_map: dict[int, str] = {}
    hostname_col: Optional[int] = None
    for idx, h in enumerate(header):
        field = _EXCEL_HEADER_MAP.get(h)
        if field:
            col_map[idx] = field
            if field == "hostname":
                hostname_col = idx

    if hostname_col is None:
        raise HTTPException(
            status_code=400,
            detail="'호스트명' 열을 찾을 수 없습니다. 엑셀 첫 행에 '호스트명' 헤더가 필요합니다.",
        )

    created = 0
    updated = 0
    skipped = 0
    errors: List[str] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        hostname_val = str(row[hostname_col]).strip() if hostname_col < len(row) and row[hostname_col] is not None else ""
        if not hostname_val:
            skipped += 1
            continue

        device_name = hostname_val
        try:
            asset = db.query(Asset).filter(Asset.device_name == device_name).first()
            is_new = asset is None
            if is_new:
                asset = Asset(device_name=device_name)
                db.add(asset)
            elif asset.deleted:
                asset.deleted = False

            for col_idx, field in col_map.items():
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx]
                if cell_val is None:
                    continue
                str_val = str(cell_val).strip()
                if not str_val:
                    continue

                if field in _AUTO_FIELD_MAP:
                    setattr(asset, _AUTO_FIELD_MAP[field], str_val)
                else:
                    setattr(asset, field, str_val)

            if is_new:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"{row_idx}행: {e}")

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "total_rows": len(rows) - 1,
    }


@router.post("/sync")
def sync_from_logs(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """최신 로그 데이터로 자동 필드를 동기화합니다."""
    snap_map = _get_latest_snaps(db)
    if not snap_map:
        return {"synced": 0, "created": 0}

    synced = 0
    created = 0
    for device_name, snap in snap_map.items():
        asset = db.query(Asset).filter(Asset.device_name == device_name).first()
        if asset and asset.deleted:
            continue
        if not asset:
            asset = Asset(device_name=device_name)
            db.add(asset)
            created += 1

        # 로그 값이 있을 때만 덮어쓰기
        field_vals = {
            "manufacturer_override": snap.manufacturer,
            "model_override": snap.system_type,
            "serial_number_override": snap.serial_number,
            "hostname_override": snap.sysname or snap.device_name,
            "os_override": snap.primary_version,
            "ip_override": snap.management_ip,
        }
        changed = False
        for col, val in field_vals.items():
            if val is not None and getattr(asset, col) != val:
                setattr(asset, col, val)
                changed = True
        if changed:
            synced += 1

    db.commit()
    return {"synced": synced, "created": created}


# ── 동적 경로 (맨 아래) ──


@router.put("/{device_name}")
def update_asset(
    device_name: str,
    data: AssetUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """특정 장비의 자산 필드를 업데이트합니다."""
    asset = db.query(Asset).filter(Asset.device_name == device_name).first()
    if not asset:
        asset = Asset(device_name=device_name)
        db.add(asset)
    elif asset.deleted:
        asset.deleted = False

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if key in _AUTO_FIELD_MAP:
            setattr(asset, _AUTO_FIELD_MAP[key], value)
        else:
            setattr(asset, key, value)

    db.commit()
    db.refresh(asset)
    return _asset_row(asset)


@router.delete("/{device_name}")
def delete_asset(
    device_name: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """특정 장비를 자산관리 목록에서 숨깁니다."""
    asset = db.query(Asset).filter(Asset.device_name == device_name).first()
    if not asset:
        asset = Asset(device_name=device_name, deleted=True)
        db.add(asset)
    else:
        asset.deleted = True

    db.commit()
    return {"deleted": True, "device_name": device_name}
